# -*- coding: utf-8 -*-
"""Build a spreadsheet the store owner can actually review prices in.

One row per product as the site now shows it, with the price it had on the
current meubleconfort.com beside it, where the new figure came from, and an
empty column to write a correction in. Anything the audit could not settle is
flagged and sorted to the top.
"""
import json, os, re, collections, statistics
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'exports', 'revision-prix.xlsx')
SITE = 'https://confort-demo.vercel.app'
CATS = {'salon': 'salon', 'chambre': 'chambre',
        'salle-a-manger': 'salle-a-manger', 'bureau': 'bureau'}

INK = '1A1A1A'; RED = 'BB3500'; PAPER = 'F4F4F3'; FLAG = 'FBE9E3'; SOFT = 'FAFAF9'

def load():
    cat = json.load(open(os.path.join(HERE, 'catalogue.json'), encoding='utf-8'))
    raw = []
    for f in ('raw1.json', 'raw2.json'):
        p = os.path.join(HERE, f)
        if os.path.exists(p):
            raw += json.load(open(p, encoding='utf-8'))['products']
    live = {}
    for x in raw:
        prices = [float(v['price']) for v in x['variants'] if v.get('price')]
        if prices:
            live[x['handle']] = min(prices)
    prices_list = {}
    p = os.path.join(HERE, 'supplier-prices.json')
    if os.path.exists(p):
        prices_list = json.load(open(p, encoding='utf-8'))
    return cat, live, prices_list

def norm(s):
    return re.sub(r'[^A-Z0-9]', '', (s or '').upper())

def flags(cat):
    """Everything the audit could not settle on its own."""
    out = collections.defaultdict(list)
    bysub = collections.defaultdict(list)
    for p in cat:
        bysub[p['sub_fr']].append(p)
    for sub, ps in bysub.items():
        if len(ps) < 4:
            continue
        med = statistics.median([x['price'] for x in ps])
        for p in ps:
            if not med:
                continue
            r = p['price'] / med
            if r > 4:
                out[id(p)].append(f'{r:.1f}× la médiane de « {sub} » ({med:,.0f} $)'.replace(',', ' '))
            elif r < 0.25:
                out[id(p)].append(f'{r:.2f}× la médiane de « {sub} » ({med:,.0f} $)'.replace(',', ' '))
    # a piece that costs more than a cheaper piece of the same suite
    fam = collections.defaultdict(dict)
    for p in cat:
        fam[re.split(r'\s+—\s+', p['name_fr'])[0].strip()][p['sub']] = p
    RANK = ['fauteuil', 'causeuse', 'canape', 'sectionnel']
    LABEL = {'fauteuil': 'fauteuil', 'causeuse': 'causeuse',
             'canape': 'canapé', 'sectionnel': 'sectionnel'}
    for b, d in fam.items():
        have = [s for s in RANK if s in d]
        for i in range(len(have) - 1):
            lo, hi = d[have[i]], d[have[i + 1]]
            if lo['price'] > hi['price']:
                out[id(hi)].append(
                    f'moins cher que le {LABEL[have[i]]} du même ensemble ({lo["price"]:,.0f} $)'.replace(',', ' '))
    return out

def main():
    cat, live, plist = load()
    fl = flags(cat)
    # A piece split out of a set never existed on its own online, so quoting the
    # parent's old price beside it would read as a price cut that never happened.
    from collections import Counter
    per_handle = Counter(x.get('handle_old') for x in cat)

    rows = []
    for p in cat:
        is_child = ' — ' in p['name_fr'] and per_handle.get(p.get('handle_old'), 0) > 1
        prev = None if is_child else live.get(p.get('handle_old'))
        if is_child:
            src = 'Nouvelle fiche — pièce d’un ensemble'
        else:
            src = ('Liste du fournisseur' if (p.get('sku') and norm(p['sku']) in plist)
                   else 'Prix boutique actuel')
        delta = (p['price'] - prev) / prev if prev else None
        note = ' ; '.join(fl.get(id(p), []))
        rows.append({
            'code': p.get('sku') or '',
            'nom': p['name_fr'],
            'cat': p['sub_fr'],
            'prix': round(p['price'], 2),
            'prev': round(prev, 2) if prev else None,
            'delta': delta,
            'mois': p.get('monthly'),
            'src': src,
            'note': note,
            'url': f"{SITE}/fr/{CATS.get(p['cat'], p['cat'])}/{p['slug']}/",
        })
    rows.sort(key=lambda r: (not r['note'], r['cat'], -r['prix']))

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Révision des prix'
    title = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor=INK)
    thin = Side(style='thin', color='DDDDDD')
    cols = [('Code', 13), ('Produit', 46), ('Catégorie', 24), ('Prix sur le site', 15),
            ('Prix actuel en ligne', 18), ('Écart', 9), ('$/mois', 8),
            ('Source du prix', 20), ('Prix corrigé', 14), ('À vérifier', 52), ('Page', 12)]
    ws.append([c[0] for c in cols])
    for i, (name, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i); c.font = title; c.fill = head_fill
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    for r in rows:
        ws.append([r['code'], r['nom'], r['cat'], r['prix'], r['prev'],
                   r['delta'], r['mois'], r['src'], None, r['note'], 'Voir'])
        i = ws.max_row
        flagged = bool(r['note'])
        for col in range(1, len(cols) + 1):
            c = ws.cell(row=i, column=col)
            c.border = Border(bottom=thin)
            if flagged:
                c.fill = PatternFill('solid', fgColor=FLAG)
            elif i % 2 == 0:
                c.fill = PatternFill('solid', fgColor=SOFT)
        ws.cell(row=i, column=4).number_format = '#,##0.00 $'
        ws.cell(row=i, column=5).number_format = '#,##0.00 $'
        ws.cell(row=i, column=6).number_format = '+0 %;-0 %;—'
        ws.cell(row=i, column=7).number_format = '#,##0 $'
        ws.cell(row=i, column=9).number_format = '#,##0.00 $'
        ws.cell(row=i, column=9).fill = PatternFill('solid', fgColor='FFFFFF')
        ws.cell(row=i, column=10).alignment = Alignment(wrap_text=True, vertical='top')
        link = ws.cell(row=i, column=11)
        link.hyperlink = r['url']; link.font = Font(color='0563C1', underline='single')

    # a short read-me sheet so the file explains itself
    info = wb.create_sheet('Comment lire ce fichier')
    info.column_dimensions['A'].width = 110
    lines = [
        ('Révision des prix — Meuble Confort & Style', True),
        ('', False),
        (f'{len(rows)} produits. Les lignes surlignées demandent une vérification.', False),
        ('', False),
        ('Prix sur le site — ce que la démo affiche aujourd’hui.', False),
        ('Prix actuel en ligne — ce que meubleconfort.com affiche présentement.', False),
        ('Écart — la différence entre les deux.', False),
        ('Source du prix — « Liste supplier 2026 » si le prix vient du fichier de prix fourni,', False),
        ('    « Prix boutique actuel » si le prix de la boutique a été conservé.', False),
        ('Prix corrigé — colonne vide : inscrivez le bon prix s’il y a lieu.', False),
        ('À vérifier — pourquoi la ligne est signalée.', False),
        ('', False),
        ('Les ensembles et leurs pièces sont maintenant séparés : l’ensemble porte le prix', False),
        ('de l’ensemble, et chaque pièce (canapé, causeuse, fauteuil…) a sa propre ligne', False),
        ('et son propre prix.', False),
        ('', False),
        ('Renvoyez le fichier avec la colonne « Prix corrigé » remplie et les prix seront', False),
        ('appliqués au site.', False),
    ]
    for text, bold in lines:
        info.append([text])
        if bold:
            info.cell(row=info.max_row, column=1).font = Font(size=14, bold=True, color=RED)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    n_flag = sum(1 for r in rows if r['note'])
    print(f'{OUT}\n  {len(rows)} products, {n_flag} flagged for review')
    for r in rows[:8]:
        if r['note']:
            print(f'   {r["nom"][:34]:34} {r["prix"]:8.2f}  {r["note"][:56]}')

if __name__ == '__main__':
    main()
